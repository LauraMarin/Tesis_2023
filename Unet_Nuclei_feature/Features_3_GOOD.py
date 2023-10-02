from __future__ import division
import cv2
from statistics import mean 
import skimage.exposure as sk_exposure
import numpy as np
import os
import statistics
import matplotlib.pyplot as plt
from skimage.measure import find_contours
from skimage import img_as_ubyte
from scipy.spatial import Delaunay
from histomicstk.preprocessing.color_normalization import reinhard
from histomicstk.preprocessing.color_conversion import lab_mean_std
import random as rng
import math
import alphashape
from skimage import color
from statistics import mean
from histmatch import hist_match
from skimage.color import separate_stains,fgx_from_rgb
from matplotlib.colors import LinearSegmentedColormap
from skimage.color import rgb2hed
from colortransfert import color_transfer
from lumen_seg_2 import *
from scipy.signal import find_peaks
from skimage.morphology import label
from skimage.io import imshow
import imutils
from scipy import ndimage
from itertools import combinations
import scipy as sp
import skimage.io
import matplotlib.pyplot as plt
from skimage.morphology import label
from skimage.morphology import watershed
from scipy import asarray as ar,exp
from scipy.spatial import Delaunay
from skimage.color import rgb2hed
from skimage.segmentation import slic
from skimage.segmentation import mark_boundaries
from statistics import stdev 
from openpyxl import Workbook
from openpyxl import load_workbook
from skimage import morphology
from skimage.feature import greycomatrix, greycoprops
from skimage.segmentation import felzenszwalb
import warnings
warnings.filterwarnings('ignore')
def get_mpl_colormap(cmap):
    

    # Initialize the matplotlib color map
    sm = plt.cm.ScalarMappable(cmap=cmap)

    # Obtain linear color range
    color_range = sm.to_rgba(np.linspace(0, 1, 256), bytes=True)[:,2::-1]

    return color_range.reshape(256, 1, 3)

def get_features_conf(indice_folder):
    file_excel="C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/output/Norm_3.xlsx"
    source_image= cv2.imread('C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/codes/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png',-1)
    meanRef, stdRef = lab_mean_std(source_image)

    list_img_final_label = []
    wb=load_workbook(file_excel)
    sheet = wb.active

   
    suffix = str(indice_folder).zfill(3)
    path_mask_Glands_= 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_glands'
    path_mask_Nuclei_rgb_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_rgb'
    path_image_RGB_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/wsi/tiles_png'
    path_final = 'C:/Users/laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/Output_final'
#for num, name_files in (enumerate(os.listdir(image_predic_path))):
 #   print(name_files)/
#name_image_mask = path_mask_Glands +  '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png '# TCGA-001-tile-r6-c110-x111622-y5122-w1024-h1024_class #_class/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'
#name_image = path_image_RGB + '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png '#'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png'
#image_nuclei =path_mask_Nuclei+'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'


    indice_row = 2
    OUTPUT_DIR = path_final + '/' +suffix
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

   
    
    str1 = ''.join(suffix)
    sheet.cell(row=indice_row, column=2).value=str(str1)
  
    path_mask_Glands = path_mask_Glands_ +'/'+suffix
  #  path_mask_Nuclei_hed=path_mask_Nuclei_hed_  +'/'+suffix
    path_mask_Nuclei_rgb=path_mask_Nuclei_rgb_  +'/'+suffix
    path_image_RGB = path_image_RGB_ +'/'+suffix

    list_glands =os.listdir(path_mask_Glands)
    list_glands = set(list_glands)
  #  list_nuclei_hed =os.listdir(path_mask_Nuclei_hed)
  #  list_nuclei_hed = set(list_nuclei_hed)
    list_nuclei_rgb =os.listdir(path_mask_Nuclei_rgb)
    list_nuclei_rgb = set(list_nuclei_rgb)
    Ratio_Glands_lumen = []
   
    
 

    Distance_between_glands = []
    disper_type = []
    for num, fname in (enumerate(os.listdir(path_image_RGB  ))):
        
        width = 364
        height = 364
        dim = (width, height)
        Final_mask_full = np.zeros((width, height), dtype=np.uint8)
                  

        
        indice_fused = 0
        indice_indivi =0
        indice_glands_dis = 0
        indice_glands_without_lumen = 0
  
        indice_nber_glands = 0



        list_ratio_non_lumen_to_contour = []
        list_ratio_Glands_lumen = []
        list_ratio_contour_2= []
        list_ratio_ROI = []
        list_ratio_contour = []
        list_ratio_density_stroma = []
        list_ratio_density_nuclei= []
        list_value_mac = []
        list_Variance_shape_glands = []
        list_space_glands = []
        list_lumen_true=[]
        list_ratio_density_nuclei_1=[]

                                      
        p= os.path.basename(fname)
        name1 = os.path.splitext(p)[0]
        
       
      #  wb.save(file_excel)
     #   matching_nuclei_hed = [s for s in list_nuclei_hed  if name1 in s]
        matching_glands =[g for g in list_glands  if name1 in g]
        matching_nuclei_rgb = [h for h in list_nuclei_rgb  if name1 in h]
        
        if len(matching_nuclei_rgb)==0 or len(matching_glands)==0    :
            continue
        elif len(matching_nuclei_rgb)!= 0 and len(matching_glands)!= 0  :
            
            width = 364
            height = 364
            dim = (width, height)
        
            name_image = path_image_RGB +'/'+  fname
            img_glands_RGB_ = cv2.imread(name_image ,-1)
            print(np.shape(img_glands_RGB_))
            
            img_glands_RGB= cv2.resize(img_glands_RGB_, dim, interpolation = cv2.INTER_AREA)
            img_glands_RGB_reinhard = reinhard (img_glands_RGB, meanRef, stdRef)
            img_hsv = cv2.cvtColor(img_glands_RGB_reinhard, cv2.COLOR_BGR2HSV)
            blank_image = np.zeros((height,width,3), np.uint8)
            


            final_lumen_mask = np.zeros((height,width), np.uint8)
          
          
            
            str3 = ''.join(matching_nuclei_rgb)
            image_nuclei_rgb = path_mask_Nuclei_rgb +'/'+ str3
            img_nuclei_mask_rgb = cv2.imread(image_nuclei_rgb,0)
            img_nuclei_mask_rgb= cv2.resize(img_nuclei_mask_rgb, dim, interpolation = cv2.INTER_AREA)
        
         #   str4 = ''.join(matching_nuclei_hed)
         #   image_nuclei_hed = path_mask_Nuclei_hed +'/'+  str4
         #   img_nuclei_mask_hed= cv2.imread(image_nuclei_hed,0)
         #   img_nuclei_mask_hed= cv2.resize(img_nuclei_mask_hed, dim, interpolation = cv2.INTER_AREA)

####### Glands ROI ######
            str2 = ''.join(matching_glands)
            name_image_mask = path_mask_Glands +'/'+ str2
            img_glands_mask = cv2.imread(name_image_mask,0)
            print(np.shape(img_glands_mask))
            print(c)
           # img_glands_mask = zoom(img_glands_mask, 2)
            img_glands_mask  = cv2.resize(img_glands_mask , dim, interpolation = cv2.INTER_AREA)

            centerLumen_X_bitwise= []
            centerLumen_Y_bitwise = []
            Contour_lumen_bitwise= []
            
            img_nuclei_mask = img_nuclei_mask_rgb #+ img_nuclei_mask_hed
            
            shape = np.shape(img_glands_RGB)
            Mask_image_rgb = np.zeros((shape[0], shape[1],3), dtype=np.uint8)
            
            Final_image_lum = np.zeros((shape[0], shape[1]), dtype=np.uint8)


            mask_rec = cv2.bitwise_not(img_glands_mask )
            try2_2 = cv2.bitwise_and(img_glands_RGB,img_glands_RGB, mask=img_glands_mask)
            img_glands_RGB_inside = cv2.cvtColor(img_glands_RGB,cv2.COLOR_BGR2RGB)
            img_inside = cv2.cvtColor(mask_rec,cv2.COLOR_BGR2RGB)
            _, mask2_inside = cv2.threshold(mask_rec,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contoursMask_inside, hierarchy_inside = cv2.findContours(mask2_inside,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            kernel_dil = np.ones((2,2), np.uint8)
            ###keep lumen
            Mask_lumen_final ,Img_final,centerLumen_X_bitwise,centerLumen_Y_bitwise,Contour_lumen_bitwise = Img_reslt_fin(Contour_lumen_bitwise,try2_2,Mask_image_rgb,contoursMask_inside,hierarchy_inside,img_glands_RGB,img_glands_RGB_reinhard,centerLumen_X_bitwise,centerLumen_Y_bitwise)
            ##### delete stroma
            
            img_slic,mask_good ,img_deconv = image_filter_equa(Img_final,img_glands_RGB_reinhard) 
            segments = slic(img_deconv,compactness=0.1, n_segments =2900,sigma = 0.5)
            superpixels = color.label2rgb(segments,img_slic , kind='avg')
            
            img_final_mask = np.zeros((height,width), np.uint8)
            img_final_NON_RGB = Img_final[:,:,0]
            for k in range(height):
                for l in range(width):
                    if (img_final_NON_RGB[k][l]!=0): # true if (image[k][l][0]==255 and image[k][l][1]==255 and image[k][l][1]==255)
                       img_final_mask[k][l]=255

            img_lum_mask = np.zeros((height,width), np.uint8)
            img_final_NON_RGB_lumen = Mask_lumen_final[:,:,0]
            for k in range(height):
                for l in range(width):
                    if (img_final_NON_RGB_lumen[k][l]!=0): # true if (image[k][l][0]==255 and image[k][l][1]==255 and image[k][l][1]==255)
                       img_lum_mask[k][l]=255

            img_hsv_ = cv2.cvtColor(Img_final, cv2.COLOR_BGR2HSV)
            img_hsv_filter = img_hsv_[:,:,0]< 162
            ph_0 = np.ones((img_glands_RGB.shape[0], img_glands_RGB.shape[1], 3), dtype='uint8')
            ph_0[:,:,0] = img_hsv_filter
            ph_0[:,:,1] = img_hsv_filter
            ph_0[:,:,2] = img_hsv_filter
            ing_final_filtered = Img_final*ph_0
          #  cv2.imshow('Distance Transform Image', ing_final_filtered)
          #  cv2.waitKey(0)
          #  cv2.imshow('',ing_final_filtered)
          #  cv2.waitKey(0)

            

            Nuclei_minus_glands = img_final_mask-img_nuclei_mask_rgb
            #cv2.imshow('',Nuclei_minus_glands)
            #cv2.waitKey(0)
            dist = cv2.distanceTransform(Nuclei_minus_glands, cv2.DIST_L2, 3)
                             
            cv2.normalize(dist, dist, 0, 1.0, cv2.NORM_MINMAX)
         #   cv2.imshow('Distance Transform Image', dist)
         #   cv2.waitKey(0)
                             
            _, dist = cv2.threshold(dist, 0.5, 1.0, cv2.THRESH_BINARY)
                             
            dist = dist.astype('uint8')
            _, mask2_dist= cv2.threshold(dist  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contours_dist, hierarchy_dist = cv2.findContours(mask2_dist,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

            _, lumen_per_contour= cv2.threshold(img_lum_mask  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contours_lumen, hierarchy_lum = cv2.findContours(lumen_per_contour,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
         #   cv2.drawContours(Img_final ,contours_dist, -1, (0,255,0), 1)
         #   cv2.imshow('',Img_final)
         #   cv2.waitKey(0)            


            img = cv2.cvtColor(img_glands_mask,cv2.COLOR_BGR2RGB)
            _, mask2 = cv2.threshold(img_glands_mask,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            res = cv2.bitwise_and(img,img, mask=mask2)
            
            hsv2bgr = cv2.cvtColor(res, cv2.COLOR_HSV2BGR)
            rgb2gray = cv2.cvtColor(hsv2bgr, cv2.COLOR_BGR2GRAY)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            #cv2.drawContours(img_glands_RGB ,contoursMask, -1, (0,255,0), 1)
           # cv2.imshow('',img_glands_RGB)
           # cv2.waitKey(0)
            
           
            Final_image_mask_added_existence = False
            if len(contours_lumen) == 0:
                Final_mask_full = Final_mask_full +Img_final
            for c_lumen in zip(contours_lumen, hierarchy_lum[0]):
                Lumen_dis_ok =  False
                temp_lumen_dist = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                cv2.fillPoly(temp_lumen_dist, pts=[c_lumen[0]], color=(255, 255, 255))
                area_lum  = cv2.contourArea(c_lumen[0])
                            
             
                if cv2.contourArea(c_lumen[0])>10:
                    for c_dist in zip(contours_dist, hierarchy_dist[0]):
                        if cv2.contourArea(c_dist[0])>2:
                            M = cv2.moments(c_dist[0])
                            cX_lumen__= int(M["m10"] / M["m00"])
                            cY_lumen__ = int(M["m01"] / M["m00"])
                            temp__dist = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                            cv2.fillPoly(temp__dist, pts=[c_dist[0]], color=(255, 255, 255))  
                            overlay_ = cv2.pointPolygonTest(c_lumen[0],(cX_lumen__,cY_lumen__),True)
                            if overlay_> 0 :
                                bitwiseXor = cv2.bitwise_or(temp_lumen_dist, temp__dist)
                                _, mask_add= cv2.threshold(bitwiseXor[:, :, 1]  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                contours_add, hierarchy_add = cv2.findContours(mask_add,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                                for c_add in zip(contours_add, hierarchy_add[0]):
                                    area_add  = cv2.contourArea(c_add[0])
                                    if area_lum /area_add>0.98 and area_lum /area_add<1.1 :
                                        Lumen_dis_ok =  True
                                        
                                       # blank_image = blank_image 
                                    else:
                                        Lumen_dis_ok =  False
                if  Lumen_dis_ok == False :
                        Final_image_mask_added_existence =  True
                                    
                      #  Final_image_mask_added = Final_image_mask_added + temp_lumen_dist[:, :, 1]
                        Final_mask_full = Final_mask_full +temp_lumen_dist[:, :, 1]
                if  Lumen_dis_ok == True :              
                        Final_lumen_mask_added = False
                        
                        evolution = []
                        callback = store_evolution_in(evolution)
                        a = equalize_adapthist(ing_final_filtered)
                        imgResult23 = img_as_float64(a)
                       # temp_lumen_dist = cv2.dilate(temp_lumen_dist, kernel_dil, iterations=1)
                        
                        ls = morphological_geodesic_active_contour(imgResult23, 250, temp_lumen_dist, threshold=0.53,balloon = 1,
                                           smoothing=1,iter_callback=callback) #0.20
                                    
                        ls = ls.astype('uint8')
                        ls = 255*ls
                        _, mask_ls= cv2.threshold(ls[:, :, 1]  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                      
                        contours_ls, hierarchy_add = cv2.findContours(mask_ls,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                        cmax_ = max(contours_ls, key = cv2.contourArea)
                     #   cv2.drawContours(ing_final_filtered ,contours_ls, -1, (0,255,0), 1)
                     #   cv2.imshow('',ing_final_filtered)
                     #   cv2.waitKey(0)
                      #  Final_image_lum =Final_image_lum +mask_ls
                        Final_mask_full = Final_mask_full + mask_ls                   
            OUTPUT_DIR_name = OUTPUT_DIR + '/' + fname 
            cv2.imwrite(OUTPUT_DIR_name, Final_mask_full)
get_features_conf(1)   
#for indice_seg in range ()

               
    
