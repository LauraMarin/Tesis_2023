from __future__ import division
import cv2
from statistics import mean 
import skimage.exposure as sk_exposure
import numpy as np
import os
import statistics
import matplotlib.pyplot as plt
from skimage.measure import find_contours
from skimage import img_as_ubyte
from scipy.spatial import Delaunay
from histomicstk.preprocessing.color_normalization import reinhard
from histomicstk.preprocessing.color_conversion import lab_mean_std
import random as rng
import math
import alphashape
from statistics import mean
from histmatch import hist_match
from skimage.color import separate_stains,fgx_from_rgb
from matplotlib.colors import LinearSegmentedColormap
from skimage.color import rgb2hed
from colortransfert import color_transfer
from lumen_seg_2 import *
from scipy.signal import find_peaks
from skimage.morphology import label
from skimage.io import imshow
import imutils
from scipy import ndimage
from itertools import combinations
import scipy as sp
import skimage.io
import matplotlib.pyplot as plt
from skimage.morphology import label
from skimage.morphology import watershed
from scipy import asarray as ar,exp
from scipy.spatial import Delaunay
from skimage.color import rgb2hed
from skimage.segmentation import slic
from skimage.segmentation import mark_boundaries
from statistics import stdev 
from openpyxl import Workbook
from openpyxl import load_workbook
from skimage import morphology
from skimage.feature import greycomatrix, greycoprops
from skimage.segmentation import felzenszwalb

import h5py
import pandas as pd

import warnings
warnings.filterwarnings('ignore')
def get_mpl_colormap(cmap):
    

    # Initialize the matplotlib color map
    sm = plt.cm.ScalarMappable(cmap=cmap)

    # Obtain linear color range
    color_range = sm.to_rgba(np.linspace(0, 1, 256), bytes=True)[:,2::-1]

    return color_range.reshape(256, 1, 3)


file_excel="C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/codes/Name_change_DFS.xlsx"
source_image= cv2.imread('C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/codes/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png',-1)
meanRef, stdRef = lab_mean_std(source_image)


wb=load_workbook(file_excel)
sheet = wb.active

indice_folder = 1
suffix = str(indice_folder).zfill(3)
path_mask_Glands_= 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_glands'
path_mask_Nuclei_hed_ ='C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_hem'
path_mask_Nuclei_rgb_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_rgb'
path_image_RGB_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/wsi/tiles_png'
path_image_lumen = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_lumen'
path_image_nuc= 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_nuclei'

#for num, name_files in (enumerate(os.listdir(image_predic_path))):
 #   print(name_files)
#name_image_mask = path_mask_Glands +  '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png '# TCGA-001-tile-r6-c110-x111622-y5122-w1024-h1024_class #_class/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'
#name_image = path_image_RGB + '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png '#'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png'
#image_nuclei =path_mask_Nuclei+'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'


indice_row = 2
dim = (364,364, height)

indice_shape = 0
table_label_list = []
table_DFS_list = []
  
#file = tables.open_file(f"{a}.h5", mode='w')
for num_, fname_ in (enumerate(os.listdir(path_image_RGB_    ))):
    
 
    
    suffix = str(fname_)
    str1 = ''.join(suffix)
    sheet.cell(row=indice_row, column=2).value=str(str1)
  
    path_mask_Glands = path_mask_Glands_ +'/'+suffix
    path_mask_Nuclei_hed=path_mask_Nuclei_hed_  +'/'+suffix
    path_mask_Nuclei_rgb=path_mask_Nuclei_rgb_  +'/'+suffix
    path_image_RGB = path_image_RGB_ +'/'+suffix
   

    list_glands =os.listdir(path_mask_Glands)
    list_glands = set(list_glands)
    list_nuclei_hed =os.listdir(path_mask_Nuclei_hed)
    list_nuclei_hed = set(list_nuclei_hed)
    list_nuclei_rgb =os.listdir(path_mask_Nuclei_rgb)
    list_nuclei_rgb = set(list_nuclei_rgb)
    Ratio_Glands_lumen = []
   
    
 

    Distance_between_glands = []
    disper_type = []


    for num, fname in (enumerate(os.listdir(path_image_RGB  ))):
        print(str(fname))
       
        
        

        
        indice_fused = 0
        indice_indivi =0
        indice_glands_dis = 0
        indice_glands_without_lumen = 0
  
        indice_nber_glands = 0



        list_ratio_non_lumen_to_contour = []
        list_ratio_Glands_lumen = []
        list_ratio_contour_2= []
        list_ratio_ROI = []
        list_ratio_contour = []
        list_ratio_density_stroma = []
        list_ratio_density_nuclei= []
        list_value_mac = []
        list_Variance_shape_glands = []
        list_space_glands = []
        list_lumen_true=[]
        list_ratio_density_nuclei_1=[]

                                      
        p= os.path.basename(fname)
        name1 = os.path.splitext(p)[0]
        
       
      #  wb.save(file_excel)
        matching_nuclei_hed = [s for s in list_nuclei_hed  if name1 in s]
        matching_glands =[g for g in list_glands  if name1 in g]
        matching_nuclei_rgb = [h for h in list_nuclei_rgb  if name1 in h]
        
        if len(matching_nuclei_rgb)==0 or len(matching_glands)==0 or len(matching_nuclei_hed)==0   :
            continue
        elif len(matching_nuclei_rgb)!= 0 and len(matching_glands)!= 0 and len(matching_nuclei_hed)!= 0  :
            
            width = 364
            height = 364
            dim = (width, height)
            column = "Num"
  
        
            name_image = path_image_RGB +'/'+  fname
            base=os.path.basename(name_image)
            name_without_ext = os.path.splitext(base)[0]
             
            print(name_without_ext)
            
            img_glands_RGB_ = cv2.imread(name_image ,-1)
            img_glands_RGB= cv2.resize(img_glands_RGB_, dim, interpolation = cv2.INTER_AREA)
            img_glands_RGB_reinhard = reinhard (img_glands_RGB, meanRef, stdRef)
            img_hsv = cv2.cvtColor(img_glands_RGB_reinhard, cv2.COLOR_BGR2HSV)
            blank_image = np.zeros((height,width,3), np.uint8)
            blank_image[:]=(139,58,98)


            final_lumen_mask = np.zeros((height,width), np.uint8)
          
          
            
            str3 = ''.join(matching_nuclei_rgb)
            image_nuclei_rgb = path_mask_Nuclei_rgb +'/'+ str3
            img_nuclei_mask_rgb = cv2.imread(image_nuclei_rgb,0)
            img_nuclei_mask_rgb= cv2.resize(img_nuclei_mask_rgb, dim, interpolation = cv2.INTER_AREA)

            str4 = ''.join(matching_nuclei_hed)
            image_nuclei_hed = path_mask_Nuclei_hed +'/'+  str4
            img_nuclei_mask_hed= cv2.imread(image_nuclei_hed,0)
            img_nuclei_mask_hed= cv2.resize(img_nuclei_mask_hed, dim, interpolation = cv2.INTER_AREA)

####### Glands ROI ######
            str2 = ''.join(matching_glands)
            name_image_mask = path_mask_Glands +'/'+ str2
            img_glands_mask = cv2.imread(name_image_mask,0)
            img_glands_mask = zoom(img_glands_mask, 2)
            img_glands_mask  = cv2.resize(img_glands_mask , dim, interpolation = cv2.INTER_AREA)

            centerLumen_X_bitwise= []
            centerLumen_Y_bitwise = []
            Contour_lumen_bitwise= []
            
            img_nuclei_mask = img_nuclei_mask_rgb + img_nuclei_mask_hed
            
            shape = np.shape(img_glands_RGB)
            Mask_image_rgb = np.zeros((shape[0], shape[1],3), dtype=np.uint8)
            
            Final_image_lum = np.zeros((shape[0], shape[1]), dtype=np.uint8)


            mask_rec = cv2.bitwise_not(img_glands_mask )
            try2_2 = cv2.bitwise_and(img_glands_RGB,img_glands_RGB, mask=img_glands_mask)
            img_glands_RGB_inside = cv2.cvtColor(img_glands_RGB,cv2.COLOR_BGR2RGB)
            img_inside = cv2.cvtColor(mask_rec,cv2.COLOR_BGR2RGB)
            _, mask2_inside = cv2.threshold(mask_rec,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contoursMask_inside, hierarchy_inside = cv2.findContours(mask2_inside,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            kernel_dil = np.ones((2,2), np.uint8)
            ###keep lumen
            Img_final,centerLumen_X_bitwise,centerLumen_Y_bitwise,Contour_lumen_bitwise = Img_reslt_fin(Contour_lumen_bitwise,try2_2,Mask_image_rgb,contoursMask_inside,hierarchy_inside,img_glands_RGB,img_glands_RGB_reinhard,centerLumen_X_bitwise,centerLumen_Y_bitwise)
            ##### delete stroma
           
            img_slic,mask_good ,img_deconv = image_filter_equa(Img_final,img_glands_RGB_reinhard) 
            segments = slic(img_deconv,compactness=0.1, n_segments =2900,sigma = 0.5)
            superpixels = color.label2rgb(segments,img_slic , kind='avg')

            
          #  img_glands_RGB = cv2.cvtColor(img_glands_RGB,cv2.COLOR_BGR2RGB)
            img = cv2.cvtColor(img_glands_mask,cv2.COLOR_BGR2RGB)
            _, mask2 = cv2.threshold(img_glands_mask,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            res = cv2.bitwise_and(img,img, mask=mask2)
            
            hsv2bgr = cv2.cvtColor(res, cv2.COLOR_HSV2BGR)
            rgb2gray = cv2.cvtColor(hsv2bgr, cv2.COLOR_BGR2GRAY)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            #cv2.drawContours(img_glands_RGB ,contoursMask, -1, (0,255,0), 1)
         #   cv2.imshow('',Img_final)
          #  cv2.waitKey(0)
            

####LUMEN####

            cmap_eosin = LinearSegmentedColormap.from_list('mycmap', ['darkviolet',
                                               'white'])


#     
            area_thr_mask=500
            name_img_lum = path_image_lumen + '/'+ fname
            name_img_nuc = path_image_nuc + '/'+ fname 
            if len(contoursMask) == 0:
                break
            else:
                for c in zip(contoursMask, hierarchy[0]):
                    indice_glands_dis_lumen = 0
                    Final_image_mask_added = np.zeros((shape[0], shape[1]), dtype=np.uint8)
                    
                    height_full_image, width_full_image= img_glands_mask.shape 
                    area_full_image = height_full_image * width_full_image 
  
                    if cv2.contourArea(c[0]) > area_thr_mask and c[1][3] == -1:
                        
        ########LUMEN#####
        
                        centerLumen_X= []
                        centerLumen_Y = []
                        area_lumen=[]
                        perimeter_lumen=[]
                        ellipse = []
                        Contour_lumen = []
                        corner_tab_0 = []
                        corner_tab_1= []
                        indice_lumen=0
                      
                    
                        contour_glands_final_no_lumen = cv2.contourArea(c[0])
                        if c[1][3] == -1 :
                            #c[1][2] != -1
                            temp = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                            cv2.fillPoly(temp, pts=[c[0]], color=(255, 255, 255))
                           
                            area_lumen,masked_image_nuclei_,Contour_lumen, masked_imageGlands_Black,imgResult,centerLumen_X,centerLumen_Y,indice_lumen_,corner_tab_0,corner_tab_1=convex_glands(Contour_lumen_bitwise,centerLumen_X_bitwise,centerLumen_Y_bitwise,kernel_dil,Img_final ,img_nuclei_mask,c,img_glands_RGB,img_glands_RGB_reinhard,centerLumen_X,centerLumen_Y,indice_lumen,area_lumen,Contour_lumen,corner_tab_0,corner_tab_1)
      ######
                        #    cv2.imshow('',masked_imageGlands_Black[:, :, 1])
                        #    cv2.waitKey(0)
                
                        
                        Nuclei_mask_separated=  masked_image_nuclei_[0]
                        contours_Nuclei_Separated = contour_thres (Nuclei_mask_separated)
                        area_thr_Nuclei = 10
                        indice_nuclei = 0
                        cX_nuclei =[]
                        cY_nuclei =[]
                        cX_corner =[]
                        cY_corner =[]
                        if len(contours_Nuclei_Separated)>0:
                            cX_nuclei,cY_nuclei,indice_nuclei  =center_nuclei( contours_Nuclei_Separated,cX_nuclei,cY_nuclei,area_thr_Nuclei,indice_nuclei )

                        
                      #  img_erosion_nuclei= cv2.erode(Nuclei_mask_separated, kernel_dil, iterations=2)
                  
                        img_dilation_nuclei= cv2.dilate(Nuclei_mask_separated, kernel_dil, iterations=5)
                       
                        ph_0 = np.ones((img_glands_RGB.shape[0], img_glands_RGB.shape[1], 3), dtype='uint8')
                        ph_0[:,:,0] = img_dilation_nuclei
                        ph_0[:,:,1] = img_dilation_nuclei
                        ph_0[:,:,2] = img_dilation_nuclei
                        nuclei_only = cv2.bitwise_and(blank_image, ph_0)
                        img_erosion_nuclei= cv2.erode(nuclei_only, kernel_dil, iterations=2)
                        
                       
                       
                       #  
                        Nuclei_minus_glands = temp[:, :, 1] -Nuclei_mask_separated
                        

                        indice_lumen = len(centerLumen_X)
                        if indice_lumen == 0  :
                            continue
                        if indice_lumen > 0  and  indice_nuclei>3 :
                             for indice_numberLumen in range(0,len(centerLumen_X)):
                                
    
                                     distance_from_lumen= []
                                     cX_nuclei_final_contour =[]
                                     cY_nuclei_final_contour =[]
                                     indice_nuclei_final_contour= 0
                                     points_lumen = []
                                     Points_convex = []
                                             
                                     points_nuclei = []
                                     point_corner= []
                                     indice_glands_nuclei = []
                  
                                     points_nuclei= orga_nuclei (points_nuclei,indice_nuclei,cX_nuclei,cY_nuclei)
                                  #   if len(corner_tab_0[indice_numberLumen]) > 3 :
                                  #       points_nuclei_2,indice_glands_nuclei = corner_ (corner_tab_1,corner_tab_0,indice_numberLumen,points_nuclei,indice_glands_nuclei)
                                  #   points_nuclei_inGland=[ points_nuclei_2[indice_Delau] for indice_Delau in indice_glands_nuclei]
                                  #   points_nuclei_inGland_,points_nuclei_inGland = del_corner (points_nuclei_inGland)
                                  #   cv = ConvexHull(points_nuclei_inGland_)
                                  #   hull_points = cv.vertices
                                  #   for indice_hull_points in hull_points :
                                  #       Points_convex.append(points_nuclei_inGland_[indice_hull_points ])
    
                                   #  Points_convex = np.array(Points_convex )
                                    
                                  #   tri = Delaunay(points_nuclei_inGland_)
                                     
                                 #    plt.imshow(img_glands_RGB)
                                 #    plt.triplot(points_nuclei_inGland_[:,0], points_nuclei_inGland_[:,1], tri.simplices)
                                 #    plt.plot(points_nuclei_inGland_[:,0], points_nuclei_inGland_[:,1], 'o')
                                 #  
                                     grid_1 = np.zeros(img_glands_RGB.shape[:2])

                                  #   cv2.fillPoly(grid_1, pts=[Points_convex], color=(255, 255, 255))
                                  #   final_lumen_mask =final_lumen_mask + grid_1 
                                     
            
               

                #   final_lumen_mask =final_lumen_mask + masked_imageGlands_Black[:, :, 1]
    
       # if indice_shape ==0 :
    # 
        final_lumen_mask =final_lumen_mask +  masked_imageGlands_Black[:, :, 1]
        final_lumen_mask_nuclei  = final_lumen_mask + Nuclei_mask_separated

        final_lumen_mask_nuclei = final_lumen_mask_nuclei.astype("uint8")                            
        _, mask2_lumen_nuc = cv2.threshold(final_lumen_mask_nuclei,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        final_lumen_rgb_nuclei_ = cv2.bitwise_and(img_glands_RGB_reinhard,img_glands_RGB_reinhard,mask= mask2_lumen_nuc)
       # height, width = final_lumen_rgb_nuclei_.shape[:2]
    #    plt.show()
     #   cv2.imshow('',final_lumen_rgb_nuclei_)
     #   cv2.waitKey(0)
        
       
       
       
        
        name_img_lum_flip_3 = path_image_lumen + '/'+ name_without_ext+ '_003.png'
        name_img_lum_flip_4 = path_image_lumen + '/'+  name_without_ext+ '_004.png'
        name_img_lum_flip_5 = path_image_lumen + '/'+ name_without_ext+ '_005.png'
        
        name_img_nuc_flip_3 = path_image_nuc + '/'+  name_without_ext+ '_003.png'
        name_img_nuc_flip_4 = path_image_nuc + '/'+  name_without_ext+ '_004.png'
        name_img_nuc_flip_5 = path_image_nuc + '/'+  name_without_ext+ '_005.png'
   
     #   rotation_matrix_nuclei = cv2.getRotationMatrix2D((width/2,height/2),45,.5)
     #   rotated_image_nuclei = cv2.warpAffine(final_lumen_rgb_nuclei_,rotation_matrix_nuclei,(width,height))

        flip_nuclei = cv2.flip(final_lumen_rgb_nuclei_,3)
        flip_nuclei_1 = cv2.flip(final_lumen_rgb_nuclei_,0)
        flip_nuclei_2 = cv2.flip(final_lumen_rgb_nuclei_,1)
  #3      flip_nuclei_3 = cv2.flip(final_lumen_rgb_nuclei_,4)
     #   cv2.imshow('',final_lumen_rgb_nuclei_)
      #  cv2.waitKey(0)
      #  cv2.imshow('',flip_nuclei)
      #  cv2.waitKey(0)

            

        _, mask2_lumen = cv2.threshold(final_lumen_mask,0, 255, cv2.THRESH_BINARY )
        final_lumen_rgb = cv2.bitwise_and(img_glands_RGB_reinhard,img_glands_RGB_reinhard,mask= mask2_lumen)
      #  rotation_matrix = cv2.getRotationMatrix2D((width/2,height/2),45,.5)
       # rotated_image_lumen = cv2.warpAffine(final_lumen_rgb,rotation_matrix,(width,height))
      #  cv2.imshow('',final_lumen_rgb)
      #  cv2.waitKey(0)
        flip_lumen = cv2.flip(final_lumen_rgb,3)
        flip_lumen_1 = cv2.flip(final_lumen_rgb,0)
        flip_lumen_2 = cv2.flip(final_lumen_rgb,1)
    #    flip_lumen_3 = cv2.flip(final_lumen_rgb,2)


        #cv2.imwrite(name_img_nuc_flip, final_lumen_rgb_nuclei_)
      
        cv2.imwrite(name_img_lum_flip_3, flip_lumen_1)
    #    cv2.imwrite(name_img_lum_flip_4, flip_lumen_2)
     #   cv2.imwrite(name_img_lum_flip_5, flip_lumen_3)

        cv2.imwrite(name_img_nuc_flip_3, flip_nuclei_1)
   #     cv2.imwrite(name_img_nuc_flip_4, flip_nuclei_2)
   #     cv2.imwrite(name_img_nuc_flip_5, flip_nuclei_3)
    #    cv2.imwrite(name_img_lum_flip, final_lumen_rgb)
    #    cv2.imwrite(name_img_lum_rot, flip_lumen)
       

   
